import io
import json
import os
import re
import sqlite3
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from xml.sax.saxutils import escape
from hmac import compare_digest
from pathlib import Path

from dotenv import load_dotenv
from flask import (
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A5
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.security import check_password_hash

from database import (
    ALLOWED_TABLES,
    BUILTIN_REGISTER_LABELS,
    GATEPASS_MUTABLE_FIELDS,
    TABLE_COLUMNS,
    TABLE_ORDER,
    all_asset_register_keys_in_order,
    allocate_next_gatepass_no,
    asset_register_column_names,
    generate_create_register_table_sql,
    get_connection,
    get_extra_register_table_row,
    is_known_asset_register_table,
    list_extra_register_table_rows,
    migrate,
    register_extra_table_exists,
    resolve_template_table_for_register,
    validate_new_register_table_key,
    normalize_excel_sheet_title,
)
from export_common import (
    EXPORT_SHEET_TITLES,
    GATEPASS_COLUMN_ORDER,
    GATEPASS_LABELS,
    column_for_excel_import_header,
    excel_value,
    header_labels_for_asset_table,
)

load_dotenv(Path(__file__).resolve().parent / ".env")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-change-me")
app.permanent_session_lifetime = timedelta(days=7)


def _auth_user():
    return (os.environ.get("ASSET_REGISTER_AUTH_USER") or "").strip()


def auth_explicitly_disabled():
    """Local/dev only: set ASSET_REGISTER_AUTH_DISABLED=1 to skip login (not for production)."""
    v = (os.environ.get("ASSET_REGISTER_AUTH_DISABLED") or "").strip().lower()
    return v in ("1", "true", "yes", "on")


def auth_configured():
    """True when username and password (or password hash) are set in the environment."""
    if not _auth_user():
        return False
    if os.environ.get("ASSET_REGISTER_AUTH_PASSWORD_HASH"):
        return True
    return bool(os.environ.get("ASSET_REGISTER_AUTH_PASSWORD"))


def auth_enforced():
    """When True, all routes require login (unless credentials missing — then only /login)."""
    return not auth_explicitly_disabled()


def _password_ok(plain: str) -> bool:
    ph = os.environ.get("ASSET_REGISTER_AUTH_PASSWORD_HASH")
    if ph:
        return check_password_hash(ph, plain or "")
    expected = os.environ.get("ASSET_REGISTER_AUTH_PASSWORD", "")
    if not expected or plain is None:
        return False
    try:
        a = plain.encode("utf-8")
        b = expected.encode("utf-8")
        if len(a) != len(b):
            return False
        return compare_digest(a, b)
    except Exception:
        return False


def _safe_next(target: str):
    if not target or not target.startswith("/") or target.startswith("//"):
        return url_for("index")
    return target


@app.context_processor
def inject_auth():
    return {"auth_enabled": auth_enforced() and auth_configured()}


@app.before_request
def require_login():
    if not auth_enforced():
        return None
    if request.endpoint in ("login", "static"):
        return None
    if request.path.startswith("/static/"):
        return None
    if not auth_configured():
        if request.path.startswith("/api/"):
            return (
                jsonify(
                    {
                        "error": (
                            "Authentication required but server is not configured: "
                            "set ASSET_REGISTER_AUTH_USER and ASSET_REGISTER_AUTH_PASSWORD "
                            "(or ASSET_REGISTER_AUTH_PASSWORD_HASH). "
                            "For local development only, set ASSET_REGISTER_AUTH_DISABLED=1."
                        )
                    }
                ),
                503,
            )
        if request.endpoint == "login":
            return None
        return redirect(url_for("login"))
    if session.get("auth_ok") is True:
        return None
    if request.path.startswith("/api/"):
        return jsonify({"error": "Unauthorized"}), 401
    return redirect(url_for("login", next=request.path))


@app.route("/login", methods=["GET", "POST"])
def login():
    if not auth_enforced():
        return redirect(url_for("index"))
    if not auth_configured():
        return render_template("login.html", error=None, next="", misconfigured=True)
    error = None
    next_url = request.args.get("next") or ""
    if request.method == "POST":
        next_url = request.form.get("next") or next_url
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        if username == _auth_user() and _password_ok(password):
            session.clear()
            session["auth_ok"] = True
            session.permanent = True
            return redirect(_safe_next(next_url))
        error = "Invalid username or password."
    return render_template("login.html", error=error, next=next_url, misconfigured=False)


@app.route("/logout")
def logout():
    session.clear()
    if auth_enforced() and auth_configured():
        return redirect(url_for("login"))
    return redirect(url_for("index"))


def _validate_table(name: str) -> str:
    conn = get_connection()
    try:
        if is_known_asset_register_table(conn, name):
            return name
    finally:
        conn.close()
    abort(404)


def _summary_register_specs(conn):
    rows = []
    for k in TABLE_ORDER:
        rows.append((EXPORT_SHEET_TITLES[k], k))
    for r in list_extra_register_table_rows(conn):
        rows.append((r["display_label"], r["table_key"]))
    rows.append((EXPORT_SHEET_TITLES["gatepass"], "gatepass"))
    return rows


def _asset_register_columns_and_headers(conn, db_key: str):
    if db_key in TABLE_COLUMNS:
        cols = TABLE_COLUMNS[db_key]
        headers = header_labels_for_asset_table(db_key)
        return cols, headers
    tpl = resolve_template_table_for_register(conn, db_key)
    if tpl is None:
        abort(404)
    cols = TABLE_COLUMNS[tpl]
    headers = header_labels_for_asset_table(tpl)
    return cols, headers


def _excel_sheet_title_for_register(conn, db_key: str) -> str | None:
    if db_key in EXPORT_SHEET_TITLES:
        return EXPORT_SHEET_TITLES[db_key]
    row = get_extra_register_table_row(conn, db_key)
    return row["excel_sheet_title"] if row else None


def _excel_sheet_title_available(conn, title: str) -> bool:
    tl = (title or "").strip().lower()
    for k in TABLE_ORDER:
        if EXPORT_SHEET_TITLES[k].lower() == tl:
            return False
    if EXPORT_SHEET_TITLES["gatepass"].lower() == tl:
        return False
    row = conn.execute(
        "SELECT table_key FROM register_extra_tables WHERE lower(excel_sheet_title) = ?",
        (tl,),
    ).fetchone()
    return row is None


def _next_unique_workbook_sheet_name(used: set[str], base: str) -> str:
    base = (base or "Sheet")[:31]
    name = base
    n = 2
    while name in used or not name:
        suffix = f" ({n})"
        name = (base[: max(1, 31 - len(suffix))] + suffix).strip()
        n += 1
    used.add(name)
    return name


def _slug_register_key_from_label(label: str) -> str:
    s = re.sub(r"[^a-z0-9_]+", "_", label.lower().strip())
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        return "register"
    if s[0].isdigit():
        s = "t_" + s
    return s[:48]


def _qi(ident: str) -> str:
    return '"' + ident.replace('"', '""') + '"'


def _qt(table: str) -> str:
    return '"' + table.replace('"', '""') + '"'


def _normalize_is_free(val):
    if val is None or val == "":
        return 0
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, int):
        return 1 if val else 0
    s = str(val).strip().lower()
    if s in ("yes", "y", "1", "true"):
        return 1
    return 0


def _row_to_dict(row):
    return {k: row[k] for k in row.keys()}


def _filter_payload_with_allowed(data: dict, allowed: set) -> dict:
    out = {}
    for key, val in data.items():
        if key not in allowed:
            continue
        if key == "is_free":
            out[key] = _normalize_is_free(val)
        elif val == "":
            out[key] = None
        else:
            out[key] = val
    return out


def _filter_payload_for_table(conn, table: str, data: dict) -> dict:
    allowed = set(asset_register_column_names(conn, table))
    return _filter_payload_with_allowed(data, allowed)


def _coerce_excel_import_cell(col: str, val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    if isinstance(val, Decimal):
        if col == "is_free":
            return int(val != 0)
        if val == val.to_integral():
            return str(int(val))
        return str(val)
    if isinstance(val, float):
        if col == "is_free":
            return int(val != 0)
        try:
            iv = int(val)
            if float(iv) == val:
                return str(iv)
        except (ValueError, OverflowError):
            pass
        return str(val)
    if isinstance(val, int):
        if col == "is_free":
            return int(val != 0)
        return str(val)
    if isinstance(val, bool):
        if col == "is_free":
            return int(val)
        return "Yes" if val else "No"
    s = str(val).strip()
    return s if s else None


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/tables", methods=["GET"])
def api_tables_list():
    conn = get_connection()
    try:
        result = []
        for t in TABLE_ORDER:
            cur = conn.execute(f"SELECT COUNT(*) AS c FROM {_qt(t)}")
            n = cur.fetchone()["c"]
            result.append(
                {
                    "name": t,
                    "row_count": n,
                    "label": BUILTIN_REGISTER_LABELS.get(t, t),
                    "columns": TABLE_COLUMNS[t],
                    "custom": False,
                }
            )
        for r in list_extra_register_table_rows(conn):
            cur = conn.execute(f"SELECT COUNT(*) AS c FROM {_qt(r['table_key'])}")
            n = cur.fetchone()["c"]
            result.append(
                {
                    "name": r["table_key"],
                    "row_count": n,
                    "label": r["display_label"],
                    "columns": TABLE_COLUMNS[r["template_table"]],
                    "custom": True,
                    "template_table": r["template_table"],
                }
            )
        gp = conn.execute("SELECT COUNT(*) AS c FROM gatepass").fetchone()["c"]
        result.append(
            {"name": "gatepass", "row_count": gp, "label": "Gatepass", "custom": False}
        )
        return jsonify(result)
    finally:
        conn.close()


@app.route("/api/tables", methods=["POST"])
def api_tables_create_register():
    data = request.get_json(silent=True) or {}
    label = (data.get("label") or "").strip()
    if not label:
        return jsonify({"error": "label is required"}), 400
    raw_key = (data.get("key") or "").strip().lower()
    key = raw_key if raw_key else _slug_register_key_from_label(label)
    ok, err = validate_new_register_table_key(key)
    if not ok:
        return jsonify({"error": err}), 400
    template = (data.get("template_table") or "laptops").strip()
    if template not in ALLOWED_TABLES:
        return jsonify({"error": "Invalid template_table"}), 400
    sheet_raw = (data.get("sheet_title") or label or key).strip()
    sheet_title = normalize_excel_sheet_title(sheet_raw, label or key)
    conn = get_connection()
    try:
        if register_extra_table_exists(conn, key):
            return jsonify({"error": "A table with this key already exists."}), 400
        if not _excel_sheet_title_available(conn, sheet_title):
            return jsonify({"error": "This Excel sheet name is already used."}), 400
        mx = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) AS m FROM register_extra_tables"
        ).fetchone()
        sort_order = int(mx["m"]) + 1
        conn.execute(
            "INSERT INTO register_extra_tables (table_key, display_label, excel_sheet_title, template_table, sort_order) "
            "VALUES (?, ?, ?, ?, ?)",
            (key, label, sheet_title, template, sort_order),
        )
        conn.execute(generate_create_register_table_sql(key, template))
        conn.commit()
        return (
            jsonify(
                {
                    "name": key,
                    "label": label,
                    "sheet_title": sheet_title,
                    "template_table": template,
                }
            ),
            201,
        )
    except sqlite3.IntegrityError:
        conn.rollback()
        return jsonify({"error": "Could not create table (duplicate name?)."}), 400
    finally:
        conn.close()


@app.route("/api/tables/<table_name>/rows", methods=["GET"])
def api_rows_list(table_name):
    t = _validate_table(table_name)
    conn = get_connection()
    try:
        cur = conn.execute(f"SELECT * FROM {_qt(t)} ORDER BY id ASC")
        rows = [_row_to_dict(r) for r in cur.fetchall()]
        return jsonify(rows)
    finally:
        conn.close()


@app.route("/api/tables/<table_name>/rows", methods=["POST"])
def api_rows_create(table_name):
    t = _validate_table(table_name)
    data = request.get_json(silent=True) or {}
    conn = get_connection()
    try:
        payload = _filter_payload_for_table(conn, t, data)
        if not payload:
            return jsonify({"error": "No valid fields"}), 400
        cols = list(payload.keys())
        placeholders = ", ".join("?" * len(cols))
        col_sql = ", ".join(_qi(c) for c in cols)
        values = [payload[c] for c in cols]
        cur = conn.execute(
            f"INSERT INTO {_qt(t)} ({col_sql}) VALUES ({placeholders})",
            values,
        )
        conn.commit()
        rid = cur.lastrowid
        cur2 = conn.execute(f"SELECT * FROM {_qt(t)} WHERE id = ?", (rid,))
        row = cur2.fetchone()
        return jsonify(_row_to_dict(row)), 201
    finally:
        conn.close()


@app.route("/api/tables/<table_name>/rows/<int:row_id>", methods=["PUT"])
def api_rows_update(table_name, row_id):
    t = _validate_table(table_name)
    data = request.get_json(silent=True) or {}
    conn = get_connection()
    try:
        payload = _filter_payload_for_table(conn, t, data)
        if not payload:
            return jsonify({"error": "No valid fields"}), 400
        cur = conn.execute(f"SELECT id FROM {_qt(t)} WHERE id = ?", (row_id,))
        if cur.fetchone() is None:
            abort(404)
        sets = ", ".join(f"{_qi(c)} = ?" for c in payload)
        values = list(payload.values()) + [row_id]
        conn.execute(f"UPDATE {_qt(t)} SET {sets} WHERE id = ?", values)
        conn.commit()
        cur2 = conn.execute(f"SELECT * FROM {_qt(t)} WHERE id = ?", (row_id,))
        return jsonify(_row_to_dict(cur2.fetchone()))
    finally:
        conn.close()


@app.route("/api/tables/<table_name>/rows/<int:row_id>", methods=["DELETE"])
def api_rows_delete(table_name, row_id):
    t = _validate_table(table_name)
    conn = get_connection()
    try:
        cur = conn.execute(f"DELETE FROM {_qt(t)} WHERE id = ?", (row_id,))
        conn.commit()
        if cur.rowcount == 0:
            abort(404)
        return "", 204
    finally:
        conn.close()


@app.route("/api/tables/<table_name>/import", methods=["POST"])
def api_rows_import_excel(table_name):
    t = _validate_table(table_name)
    f = request.files.get("file")
    if f is None or f.filename is None or f.filename.strip() == "":
        return jsonify({"error": "Missing file (use multipart field name 'file')"}), 400
    data = f.read()
    if not data:
        return jsonify({"error": "Empty file"}), 400
    wb = None
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as ex:
        return jsonify({"error": f"Invalid Excel file: {ex}"}), 400
    conn = get_connection()
    try:
        try:
            expected = _excel_sheet_title_for_register(conn, t)
            if expected and expected in wb.sheetnames:
                ws = wb[expected]
                sheet_used = expected
            else:
                ws = wb[wb.sheetnames[0]]
                sheet_used = ws.title
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return jsonify({"error": "Worksheet is empty"}), 400
            cols, hdr_labels = _asset_register_columns_and_headers(conn, t)
            col_indices = []
            for idx, cell in enumerate(header_row):
                cname = column_for_excel_import_header(cell, cols, hdr_labels)
                if cname:
                    col_indices.append((idx, cname))
            if not col_indices:
                return jsonify({"error": "No recognized column headers in the first row"}), 400

            imported = 0
            skipped_empty = 0
            errors = []
            for row_num, row in enumerate(rows_iter, start=2):
                raw = {}
                for idx, col_name in col_indices:
                    v = row[idx] if row is not None and idx < len(row) else None
                    coerced = _coerce_excel_import_cell(col_name, v)
                    if coerced is not None:
                        raw[col_name] = coerced
                payload = _filter_payload_for_table(conn, t, raw)
                if not payload:
                    skipped_empty += 1
                    continue
                cols_ins = list(payload.keys())
                placeholders = ", ".join("?" * len(cols_ins))
                col_sql = ", ".join(_qi(c) for c in cols_ins)
                values = [payload[c] for c in cols_ins]
                try:
                    conn.execute(
                        f"INSERT INTO {_qt(t)} ({col_sql}) VALUES ({placeholders})",
                        values,
                    )
                    conn.commit()
                    imported += 1
                except Exception as ex:
                    conn.rollback()
                    errors.append({"row": row_num, "error": str(ex)})

            return jsonify(
                {
                    "imported": imported,
                    "skipped_empty": skipped_empty,
                    "errors": errors,
                    "sheet_used": sheet_used,
                }
            )
        finally:
            conn.close()
    finally:
        if wb is not None:
            wb.close()


# --- Excel export (openpyxl) ---
_EXCEL_HEADER_FILL = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
_EXCEL_HEADER_FONT = Font(color="FFFFFF", bold=True)
_EXCEL_ALT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")


def _autofit_worksheet(ws, max_width=55):
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        letter = get_column_letter(col[0].column)
        maxlen = 10
        for cell in col:
            if cell.value is None:
                continue
            maxlen = max(maxlen, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = maxlen


def _style_header_row(ws, row_idx=1):
    for cell in ws[row_idx]:
        cell.fill = _EXCEL_HEADER_FILL
        cell.font = _EXCEL_HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _apply_alternating_rows(ws, start_row=2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        fill = _EXCEL_ALT_FILL if i % 2 == 1 else None
        if fill:
            for c in row:
                c.fill = fill


def _last_updated_for_table(conn, db_key: str):
    if db_key == "gatepass":
        q = "SELECT MAX(COALESCE(updated_at, created_at)) AS m FROM gatepass"
    else:
        q = f"SELECT MAX(created_at) AS m FROM {_qt(db_key)}"
    try:
        row = conn.execute(q).fetchone()
        v = row["m"] if row else None
        return v if v is not None else ""
    except Exception:
        return ""


def _write_summary_sheet(ws, conn):
    ws.append(["Table Name", "Total Rows", "Last Updated"])
    _style_header_row(ws, 1)
    for display_name, db_key in _summary_register_specs(conn):
        if db_key == "gatepass":
            cnt = conn.execute("SELECT COUNT(*) AS c FROM gatepass").fetchone()["c"]
        else:
            cnt = conn.execute(f"SELECT COUNT(*) AS c FROM {_qt(db_key)}").fetchone()["c"]
        lu = _last_updated_for_table(conn, db_key)
        ws.append([display_name, cnt, lu])
    _autofit_worksheet(ws)


def _write_asset_sheet(ws, conn, db_key: str):
    cols, headers = _asset_register_columns_and_headers(conn, db_key)
    ws.append(headers)
    _style_header_row(ws, 1)
    cur = conn.execute(f"SELECT * FROM {_qt(db_key)} ORDER BY id ASC")
    for r in cur.fetchall():
        ws.append([excel_value(c, r[c]) for c in cols])
    _apply_alternating_rows(ws, 2)
    _autofit_worksheet(ws)


def _write_gatepass_data_sheet(ws, conn):
    headers = [GATEPASS_LABELS[c] for c in GATEPASS_COLUMN_ORDER]
    ws.append(headers)
    _style_header_row(ws, 1)
    cur = conn.execute("SELECT * FROM gatepass ORDER BY id ASC")
    for r in cur.fetchall():
        d = _row_to_dict(r)
        ws.append([d.get(c) for c in GATEPASS_COLUMN_ORDER])
    _apply_alternating_rows(ws, 2)
    _autofit_worksheet(ws)


@app.route("/api/export/all")
def export_all():
    conn = get_connection()
    try:
        try:
            wb = Workbook()
            wb.remove(wb.active)
            used_names: set[str] = set()
            sum_title = _next_unique_workbook_sheet_name(used_names, "Summary")
            ws_sum = wb.create_sheet(sum_title, 0)
            _write_summary_sheet(ws_sum, conn)
            for db_key in all_asset_register_keys_in_order(conn):
                base = _excel_sheet_title_for_register(conn, db_key) or db_key
                title = _next_unique_workbook_sheet_name(used_names, base)
                ws = wb.create_sheet(title)
                _write_asset_sheet(ws, conn, db_key)
            gp_base = EXPORT_SHEET_TITLES["gatepass"]
            gp_title = _next_unique_workbook_sheet_name(used_names, gp_base)
            ws_g = wb.create_sheet(gp_title)
            _write_gatepass_data_sheet(ws_g, conn)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            fname = f"asset_register_export_{datetime.now().date().isoformat()}.xlsx"
            return send_file(
                bio,
                as_attachment=True,
                download_name=fname,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as ex:
            app.logger.exception("export_all failed")
            return jsonify({"error": str(ex)}), 500
    finally:
        conn.close()


# --- Gatepass API ---
def _filter_gatepass_payload(data: dict) -> dict:
    out = {}
    for k, v in data.items():
        if k not in GATEPASS_MUTABLE_FIELDS:
            continue
        if k == "asset_items":
            if v is None or v == "":
                out[k] = None
            elif isinstance(v, (list, dict)):
                out[k] = json.dumps(v)
            else:
                s = str(v).strip()
                out[k] = s if s else None
            continue
        out[k] = None if v == "" else v
    return out


def _fmt_gatepass_pdf_date(s):
    if not s:
        return ""
    try:
        if len(str(s)) >= 10:
            d = datetime.strptime(str(s)[:10], "%Y-%m-%d")
            return d.strftime("%d/%m/%Y")
    except Exception:
        pass
    return str(s)


def _pdf_gatepass_buffer(row: dict):
    """A5 Platypus PDF matching InfoDesk physical gate pass layout."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A5,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    company_style = ParagraphStyle(
        "gp_company",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=13,
        alignment=TA_CENTER,
        spaceAfter=2,
    )
    addr_style = ParagraphStyle(
        "gp_addr",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        alignment=TA_CENTER,
        spaceAfter=2,
        textColor=colors.HexColor("#444444"),
    )
    story.append(Paragraph("InfoDesk India Private Limited", company_style))
    story.append(Paragraph("12B Nutan Bharat Alkapuri, Vadodara", addr_style))
    story.append(
        Paragraph(
            "390007 Gujarat, India",
            ParagraphStyle("addr2", parent=addr_style, spaceAfter=8),
        )
    )

    gate_pass_label = Table(
        [["GATE PASS"]],
        colWidths=[7 * cm],
    )
    gate_pass_label.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 16),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 2, colors.black),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    wrapper = Table([[gate_pass_label]], colWidths=[doc.width])
    wrapper.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    story.append(wrapper)
    story.append(Spacer(1, 0.35 * cm))

    pass_type = str(row.get("pass_type") or "Returnable / Outward").strip() or "Returnable / Outward"
    type_style = ParagraphStyle(
        "gp_type",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    story.append(Paragraph(escape(pass_type), type_style))
    story.append(Spacer(1, 0.2 * cm))

    field_style = ParagraphStyle(
        "gp_field",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        alignment=TA_LEFT,
    )
    issued_to = escape(str(row.get("issued_to") or ""))
    person = escape(str(row.get("person") or ""))
    issued_table = Table(
        [
            [
                Paragraph(f"<b>Issued to:</b> {issued_to}", field_style),
                Paragraph(f"<b>Person:</b> {person}", field_style),
            ]
        ],
        colWidths=[doc.width * 0.58, doc.width * 0.42],
    )
    issued_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(issued_table)
    story.append(Spacer(1, 0.2 * cm))

    gp_no = escape(str(row.get("gatepass_no") or ""))
    gp_date = escape(_fmt_gatepass_pdf_date(row.get("gatepass_date")))
    no_date_table = Table(
        [
            [
                Paragraph(f"<b>No:</b> {gp_no}", field_style),
                Paragraph(f"<b>Date:</b> {gp_date}", field_style),
            ]
        ],
        colWidths=[doc.width * 0.5, doc.width * 0.5],
    )
    no_date_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(no_date_table)
    story.append(Spacer(1, 0.35 * cm))

    raw_items = row.get("asset_items") or "[]"
    try:
        if isinstance(raw_items, str):
            items = json.loads(raw_items) if raw_items.strip() else []
        elif isinstance(raw_items, list):
            items = raw_items
        else:
            items = []
    except Exception:
        items = []

    num_item_rows = 5
    cell_style = ParagraphStyle(
        "gp_cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        alignment=TA_CENTER,
        leading=9,
    )
    desc_cell_style = ParagraphStyle(
        "gp_desc",
        parent=cell_style,
        alignment=TA_LEFT,
    )
    hdr_style = ParagraphStyle(
        "gp_hdr",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        alignment=TA_CENTER,
        leading=9,
    )
    items_header = [
        [
            Paragraph("Sr.<br/>No.", hdr_style),
            Paragraph("Description", hdr_style),
            Paragraph("Unit", hdr_style),
            Paragraph("Qty.", hdr_style),
            Paragraph("Rmks", hdr_style),
        ]
    ]
    items_rows = []
    for i in range(num_item_rows):
        if i < len(items):
            it = items[i] or {}
            d = escape(str(it.get("description") or ""))
            u = escape(str(it.get("unit") or ""))
            q = escape(str(it.get("qty") or ""))
            rmk = escape(str(it.get("remarks") or ""))
            items_rows.append(
                [
                    Paragraph(str(i + 1), cell_style),
                    Paragraph(d, desc_cell_style),
                    Paragraph(u, cell_style),
                    Paragraph(q, cell_style),
                    Paragraph(rmk, cell_style),
                ]
            )
        else:
            items_rows.append(
                [
                    Paragraph(str(i + 1), cell_style),
                    Paragraph("", desc_cell_style),
                    Paragraph("", cell_style),
                    Paragraph("", cell_style),
                    Paragraph("", cell_style),
                ]
            )

    cw0 = doc.width * 0.1
    cw1 = doc.width * 0.42
    cw2 = doc.width * 0.14
    cw3 = doc.width * 0.12
    cw4 = doc.width - cw0 - cw1 - cw2 - cw3
    full_items_data = items_header + items_rows
    row_heights = [0.75 * cm] + [0.85 * cm] * num_item_rows
    items_table = Table(
        full_items_data,
        colWidths=[cw0, cw1, cw2, cw3, cw4],
        rowHeights=row_heights,
    )
    items_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
            ]
        )
    )
    story.append(items_table)
    story.append(Spacer(1, 0.45 * cm))

    sig_style = ParagraphStyle(
        "gp_sig",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        alignment=TA_CENTER,
    )
    sig_bold = ParagraphStyle(
        "gp_sig_b",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        alignment=TA_CENTER,
    )
    dh = escape(str(row.get("department_head") or ""))
    si = escape(str(row.get("security_incharge") or ""))
    sig_table = Table(
        [
            [
                Paragraph("<b>Department Head</b>", sig_bold),
                Paragraph("<b>Security Incharge</b>", sig_bold),
            ],
            [
                Paragraph(dh, sig_style),
                Paragraph(si, sig_style),
            ],
            [
                Paragraph("___________________", sig_style),
                Paragraph("___________________", sig_style),
            ],
        ],
        colWidths=[doc.width * 0.5, doc.width * 0.5],
    )
    sig_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(sig_table)
    story.append(Spacer(1, 0.35 * cm))

    recv_name = escape(str(row.get("receiver_name") or ""))
    recv_left = ParagraphStyle(
        "gp_recv",
        parent=sig_style,
        alignment=TA_LEFT,
    )
    recv_bold_left = ParagraphStyle(
        "gp_recv_b",
        parent=sig_bold,
        alignment=TA_LEFT,
    )
    recv_table = Table(
        [
            [Paragraph("<b>Receiver's Sign</b>", recv_bold_left), ""],
            [Paragraph(recv_name, recv_left), ""],
            [Paragraph("___________________", recv_left), ""],
        ],
        colWidths=[doc.width * 0.65, doc.width * 0.35],
    )
    recv_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(recv_table)

    doc.build(story)
    buf.seek(0)
    return buf


def _gatepass_s3_bucket_and_prefix():
    bucket = (
        os.environ.get("GATEPASS_S3_BUCKET") or os.environ.get("BACKUP_BUCKET") or ""
    ).strip()
    prefix = (os.environ.get("GATEPASS_S3_PREFIX") or "GatePass").strip().strip("/")
    return bucket, prefix


def _safe_s3_filename_part(text: str) -> str:
    t = re.sub(r"[^A-Za-z0-9._-]+", "_", str(text or "")).strip("_")
    return t or "unknown"


def _save_gatepass_pdf_to_s3(row_dict: dict) -> None:
    """Upload gatepass PDF under GatePass/ when S3 bucket env is set (EC2 + IAM)."""
    bucket, prefix = _gatepass_s3_bucket_and_prefix()
    if not bucket:
        return
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError:
        app.logger.warning("boto3 not installed; skipping gatepass S3 upload")
        return
    try:
        d = dict(row_dict)
        pdf_buf = _pdf_gatepass_buffer(d)
        body = pdf_buf.getvalue()
        rid = d.get("id") or 0
        gno = _safe_s3_filename_part(d.get("gatepass_no"))
        key = f"{prefix}/{rid}_{gno}.pdf"
        boto3.client("s3").put_object(
            Bucket=bucket,
            Key=key,
            Body=body,
            ContentType="application/pdf",
        )
        app.logger.info("Gatepass PDF stored at s3://%s/%s", bucket, key)
    except ClientError:
        app.logger.exception(
            "S3 PutObject failed for gatepass id=%s", row_dict.get("id")
        )
    except Exception:
        app.logger.exception(
            "Gatepass S3 upload error for id=%s", row_dict.get("id")
        )


@app.route("/api/gatepass", methods=["GET"])
def gatepass_list():
    conn = get_connection()
    try:
        cur = conn.execute("SELECT * FROM gatepass ORDER BY id DESC")
        return jsonify([_row_to_dict(r) for r in cur.fetchall()])
    finally:
        conn.close()


@app.route("/api/gatepass", methods=["POST"])
def gatepass_create():
    data = request.get_json(silent=True) or {}
    payload = _filter_gatepass_payload(data)
    conn = get_connection()
    try:
        conn.execute("BEGIN IMMEDIATE")
        try:
            gno = allocate_next_gatepass_no(conn)
            cols = ["gatepass_no"] + list(payload.keys())
            values = [gno] + [payload[k] for k in payload.keys()]
            ph = ", ".join("?" * len(values))
            colsql = ", ".join(_qi(c) for c in cols)
            cur = conn.execute(f"INSERT INTO gatepass ({colsql}) VALUES ({ph})", values)
            conn.commit()
            rid = cur.lastrowid
        except Exception:
            conn.rollback()
            raise
        row = conn.execute("SELECT * FROM gatepass WHERE id = ?", (rid,)).fetchone()
        d = _row_to_dict(row)
        _save_gatepass_pdf_to_s3(d)
        return jsonify(d), 201
    finally:
        conn.close()


@app.route("/api/gatepass/<int:row_id>", methods=["PUT"])
def gatepass_update(row_id):
    data = request.get_json(silent=True) or {}
    payload = _filter_gatepass_payload(data)
    if not payload:
        return jsonify({"error": "No valid fields"}), 400
    conn = get_connection()
    try:
        if conn.execute("SELECT id FROM gatepass WHERE id = ?", (row_id,)).fetchone() is None:
            abort(404)
        sets = ", ".join(f"{_qi(k)} = ?" for k in payload) + ", updated_at = datetime('now')"
        vals = list(payload.values()) + [row_id]
        conn.execute(f"UPDATE gatepass SET {sets} WHERE id = ?", vals)
        conn.commit()
        row = conn.execute("SELECT * FROM gatepass WHERE id = ?", (row_id,)).fetchone()
        d = _row_to_dict(row)
        _save_gatepass_pdf_to_s3(d)
        return jsonify(d)
    finally:
        conn.close()


@app.route("/api/gatepass/<int:row_id>", methods=["DELETE"])
def gatepass_delete(row_id):
    conn = get_connection()
    try:
        cur = conn.execute("DELETE FROM gatepass WHERE id = ?", (row_id,))
        conn.commit()
        if cur.rowcount == 0:
            abort(404)
        return "", 204
    finally:
        conn.close()


@app.route("/api/gatepass/<int:row_id>/pdf")
def gatepass_pdf(row_id):
    conn = get_connection()
    try:
        row = conn.execute("SELECT * FROM gatepass WHERE id = ?", (row_id,)).fetchone()
        if not row:
            abort(404)
        d = _row_to_dict(row)
        pdf = _pdf_gatepass_buffer(d)
        safe = str(d.get("gatepass_no") or row_id).replace("/", "-")
        return send_file(
            pdf,
            as_attachment=True,
            download_name=f"gatepass_{safe}.pdf",
            mimetype="application/pdf",
        )
    finally:
        conn.close()


@app.route("/api/gatepass/export")
def gatepass_export_excel():
    conn = get_connection()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Gatepass"
        _write_gatepass_data_sheet(ws, conn)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"gatepass_export_{datetime.now().date().isoformat()}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        conn.close()


with app.app_context():
    migrate()


if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(debug=debug, host="127.0.0.1", port=int(os.environ.get("PORT", "5000")))
